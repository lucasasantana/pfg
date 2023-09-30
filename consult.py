import os
from db import fetch_one, fetch_all
from openpyxl import Workbook, load_workbook

def get_file():
    excel_file_name = './results/data/results.xlsx'

    # Check if the Excel file already exists
    if os.path.exists(excel_file_name):
        # If it exists, load the existing workbook
        return load_workbook(excel_file_name)
    else:
        # If it doesn't exist, create a new workbook
        return Workbook()

def save(workbook):
    excel_file_name = './results/data/results.xlsx'
    workbook.save(excel_file_name)

def create_or_load_sheet(sheet_name, workbook):
    # Check if the 'flag_rate' sheet already exists
    if sheet_name in workbook.sheetnames:
        # If it exists, get a reference to the sheet
        workbook[sheet_name].delete_rows(1, workbook[sheet_name].max_row)
        return workbook[sheet_name]
    else:
        # If it doesn't exist, create a new sheet called 'flag_rate'
        return  workbook.create_sheet(title=sheet_name)
    
     

    
def get_all_data():
    # Excel file name
    workbook = get_file()
    sheet = create_or_load_sheet('data', workbook)

    # Query to calculate the percentage of files with feature flags
    query_sql = "SELECT * FROM flags "
    (results, descriptions) = fetch_all(query_sql)

    # Add the result to the sheet
    columns = [description[0] for description in descriptions]
    sheet.append(columns)

    # Adicionar os resultados à planilha
    for result in results:
        sheet.append(result)

    # Save the Excel file
    save(workbook)
    

def calculate_percentage_feature_flags():
    # Excel file name
    workbook = get_file()
    sheet = create_or_load_sheet('flag_rate', workbook)

    # Query to calculate the percentage of files with feature flags
    query_sql = """
    SELECT 
        (COUNT(DISTINCT file_name) * 100.0 / (SELECT COUNT(DISTINCT file_name) FROM flags)) AS percentage
    FROM flags
    WHERE number_of_flags > 0;
    """
    result = fetch_one(query_sql)

    # Add the result to the sheet
    sheet.append(['Percentage of Files with Feature Flags'])
    sheet.append([result[0]])

    # Save the Excel file
    save(workbook)

def calculate_smells():
    # Excel file name
    workbook = get_file()
    sheet = create_or_load_sheet('smells_rate', workbook)

    # Query to calculate the percentage of files with feature flags
    query_sql = """
    SELECT 
        CASE WHEN number_of_flags > 0 THEN 'Possui flags' ELSE 'Não possui flags' END AS flag_category,
        COUNT(DISTINCT file_name) AS file_count,
        SUM(blocker + critical + info + major + minor) AS total_code_smells
    FROM flags
    GROUP BY flag_category;
    """
    (results, _) = fetch_all(query_sql)

    # Add the result to the sheet
    sheet.append(['Flag Status', 'Quantidade de Arquivos', 'Total Code Smells'])
    for result in results:
        sheet.append(result)

    # Save the Excel file
    save(workbook)

def calculate_smells_by_severity():
    # Excel file name
    workbook = get_file()
    sheet = create_or_load_sheet('smells_by_severity', workbook)

    # Query to calculate the percentage of files with feature flags
    query_sql = """
    SELECT 
        CASE WHEN number_of_flags > 0 THEN 'Possui flags' ELSE 'Não possui flags' END AS flag_category,
        COUNT(DISTINCT file_name) AS distinct_file_count,
        SUM(blocker) AS blocker_count,
        SUM(critical) AS critical_count,
        SUM(info) AS info_count,
        SUM(major) AS major_count,
        SUM(minor) AS minor_count
    FROM flags
    GROUP BY flag_category;
    """
    (results, _) = fetch_all(query_sql)

    # Add the result to the sheet
    sheet.append(['Flag Status', 'Quantidade de Arquivos', 'blocker', 'critical', 'info', 'major', 'minor'])
    for result in results:
        sheet.append(result)

    # Save the Excel file
    save(workbook)

get_all_data()
calculate_percentage_feature_flags()
calculate_smells()
calculate_smells_by_severity()